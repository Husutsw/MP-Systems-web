# MP-Systems version web (Streamlit)
import streamlit as st
from datetime import datetime
import pandas as pd
from io import BytesIO
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

# Configuracion de la pagina

st.set_page_config(page_title="MP-Systems Web", layout="wide")
st.title("💼 MP-Systems - Sistema Contable en la Nube")

# Menu lateral
opciones = [
    "Inicio",
    "Libro Diario",
    "Catalogo de Cuentas",
    "Estado de Resultados",
    "Balance General",
    "Historial del Sistema"
]
seleccion = st.sidebar.selectbox("Selecciona un modulo", opciones)

# Modulo: INICIO
if seleccion == "Inicio":
    st.subheader("🏁 Configuracion Inicial del Sistema")

    with st.form(key="form_config"):
        st.write("🔧 Complete los siguientes datos para comenzar:")

        empresa = st.text_input("Nombre de la empresa", placeholder="Ejemplo: Balol Corp.")
        formato_fecha = st.selectbox("Formato de Fecha", ["dd/mm/yyyy", "mm/dd/yyyy", "yyyy-mm-dd"])
        moneda_opciones = ["RD$", "US$", "EUR$", "Otros"]
        moneda = st.selectbox("Simbolo de Moneda", moneda_opciones)

        moneda_personalizada = ""
        if moneda == "Otros":
            moneda_personalizada = st.text_input("Ingrese el simbolo personalizado de la moneda")

        boton_guardar = st.form_submit_button("Guardar configuracion")

    if boton_guardar:
        simbolo_final = moneda_personalizada if moneda == "Otros" else moneda
        st.success(f"✅ Configuracion guardada: {empresa} | {formato_fecha} | {simbolo_final}")
        st.session_state["empresa"] = empresa
        st.session_state["formato_fecha"] = formato_fecha
        st.session_state["simbolo_moneda"] = simbolo_final

# Modulo: LIBRO DIARIO
elif seleccion == "Libro Diario":
    # IMPORTACIONES NECESARIAS
    from io import BytesIO
    import pandas as pd
    from reportlab.lib.pagesizes import letter
    from reportlab.pdfgen import canvas
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    empresa = st.session_state.get("empresa", "Empresa Desconocida")
    moneda = st.session_state.get("simbolo_moneda", "RD$")
    formato_fecha = st.session_state.get("formato_fecha", "dd/mm/yyyy")

    st.subheader(f"📒 Libro Diario - {empresa}")
    st.write(f"Formato de fecha: {formato_fecha} | Moneda: {moneda}")
    st.markdown("### ➕ Ingresar Nueva Transaccion")

    with st.form("form_libro_diario"):
        col1, col2, col3 = st.columns([1, 1, 2])
        with col1:
            fecha = st.text_input("Fecha", placeholder="Ej: 19/12/2024")
        with col2:
            cuenta = st.text_input("Cuenta", placeholder="Ej: 101")
        with col3:
            descripcion = st.text_input("Descripcion", placeholder="Ej: Caja general")

        col4, col5 = st.columns(2)
        with col4:
            debe = st.text_input("Debe", placeholder="0.00")
        with col5:
            haber = st.text_input("Haber", placeholder="0.00")

        submit = st.form_submit_button("Agregar")

    if "asientos" not in st.session_state:
        st.session_state["asientos"] = []

    if submit:
        errores = []
        if not fecha.replace("/", "").isdigit():
            errores.append("La fecha debe contener solo numeros y '/'.")
        if not cuenta.strip():
            errores.append("La cuenta no puede estar vacia.")
        if not descripcion.strip():
            errores.append("La descripcion no puede estar vacia.")
        try:
            valor_debe = float(debe) if debe.strip() else 0.0
        except:
            errores.append("Debe contiene un valor invalido.")
            valor_debe = 0.0
        try:
            valor_haber = float(haber) if haber.strip() else 0.0
        except:
            errores.append("Haber contiene un valor invalido.")
            valor_haber = 0.0

        if errores:
            st.error("⚠️ Errores:\n- " + "\n- ".join(errores))
        else:
            st.session_state["asientos"].append({
                "fecha": fecha,
                "cuenta": cuenta,
                "descripcion": descripcion,
                "debe": valor_debe,
                "haber": valor_haber
            })
            st.success("✅ Transaccion agregada.")

    if st.session_state["asientos"]:
        st.markdown("### 🧾 Asientos Registrados")

        if "confirmar_eliminacion" not in st.session_state:
            st.session_state["confirmar_eliminacion"] = False

        if st.button("🧼 Eliminar Todo"):
            st.session_state["confirmar_eliminacion"] = True

        if st.session_state["confirmar_eliminacion"]:
            confirm = st.radio("¿Seguro que deseas borrar todos los registros?", ["No", "Si"], index=0)
            if confirm == "Si":
                st.session_state["asientos"].clear()
                st.session_state["confirmar_eliminacion"] = False
                st.success("✅ Asientos eliminados.")
                st.rerun()

        total_debe = 0
        total_haber = 0
        for i, asiento in enumerate(st.session_state["asientos"]):
            col1, col2 = st.columns([6, 1])
            with col1:
                st.write(f"{i+1}. {asiento['fecha']} | {asiento['cuenta']} | {asiento['descripcion']} | {moneda} {asiento['debe']:,.2f} | {moneda} {asiento['haber']:,.2f}")
            with col2:
                if st.button("🗑️", key=f"eliminar_{i}"):
                    st.session_state["asientos"].pop(i)
                    st.rerun()
            total_debe += asiento["debe"]
            total_haber += asiento["haber"]

        st.markdown("---")
        st.write(f"🔢 **Total Debe:** {moneda} {total_debe:,.2f}")
        st.write(f"🔢 **Total Haber:** {moneda} {total_haber:,.2f}")
        if total_debe == total_haber:
            st.success("✅ Asiento Cuadrado.")
        else:
            st.warning("⚠️ Asiento no cuadrado.")

        # EXCEL EXPORT CON ESTILO
        df = pd.DataFrame(st.session_state["asientos"])
        total_row = {
            "fecha": "",
            "cuenta": "",
            "descripcion": "TOTALES",
            "debe": total_debe,
            "haber": total_haber
        }
        df = pd.concat([df, pd.DataFrame([total_row])], ignore_index=True)

        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Libro Diario')
            wb = writer.book
            ws = writer.sheets['Libro Diario']
            ultima = len(df) + 1
            verde = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            rojo = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            moneda_col = [get_column_letter(4), get_column_letter(5)]
            for col in moneda_col:
                for row in range(2, ultima + 1):
                    ws[f"{col}{row}"].number_format = f'"{moneda}" #,##0.00'
            for col in moneda_col:
                ws[f"{col}{ultima}"].fill = verde if total_debe == total_haber else rojo
                ws[f"{col}{ultima}"].font = Font(bold=True)

        output.seek(0)
        st.download_button(
            label="📁 Exportar a Excel",
            data=output,
            file_name="libro_diario.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        # PDF EXPORT
        def generar_pdf(asientos, empresa, total_debe, total_haber, moneda):
            buffer = BytesIO()
            c = canvas.Canvas(buffer, pagesize=letter)
            c.setFont("Helvetica", 10)
            ancho, alto = letter
            margen = 40
            y = alto - 50

            c.drawString(margen, y, f"MP-Systems - Libro Diario")
            y -= 15
            c.drawString(margen, y, f"Empresa: {empresa}")
            y -= 15
            c.drawString(margen, y, f"Fecha del reporte: {datetime.now().strftime('%d/%m/%Y')}")
            y -= 25

            c.setFont("Helvetica-Bold", 9)
            c.drawString(margen, y, "Fecha")
            c.drawString(margen + 70, y, "Cuenta")
            c.drawString(margen + 140, y, "Descripcion")
            c.drawString(margen + 330, y, "Debe")
            c.drawString(margen + 400, y, "Haber")
            y -= 15
            c.setFont("Helvetica", 9)

            for asiento in asientos:
                if y < 50:
                    c.showPage()
                    y = alto - 50
                c.drawString(margen, y, asiento['fecha'])
                c.drawString(margen + 70, y, asiento['cuenta'])
                c.drawString(margen + 140, y, asiento['descripcion'][:30])
                c.drawRightString(margen + 390, y, f"{moneda} {asiento['debe']:,.2f}")
                c.drawRightString(margen + 460, y, f"{moneda} {asiento['haber']:,.2f}")
                y -= 15

            y -= 10
            c.setFont("Helvetica-Bold", 9)
            c.drawString(margen + 140, y, "TOTALES:")
            c.drawRightString(margen + 390, y, f"{moneda} {total_debe:,.2f}")
            c.drawRightString(margen + 460, y, f"{moneda} {total_haber:,.2f}")
            c.showPage()
            c.save()
            buffer.seek(0)
            return buffer

        pdf_buffer = generar_pdf(st.session_state["asientos"], empresa, total_debe, total_haber, moneda)
        st.download_button(
            label="📄 Exportar a PDF",
            data=pdf_buffer,
            file_name="libro_diario.pdf",
            mime="application/pdf"
        )
    else:
        st.info("No se han ingresado transacciones aun.")


# Modulo: CATALOGO DE CUENTAS
elif seleccion == "Catalogo de Cuentas":
    # IMPORTACIONES NECESARIAS
    import pandas as pd
    from io import BytesIO
    from datetime import datetime
    from reportlab.lib.pagesizes import letter
    from reportlab.pdfgen import canvas

    # RECUPERAR DATOS DE CONFIGURACION
    empresa = st.session_state.get("empresa", "Empresa Desconocida")
    moneda = st.session_state.get("simbolo_moneda", "RD$")
    formato_fecha = st.session_state.get("formato_fecha", "dd/mm/yyyy")

    # TITULO
    st.subheader(f"📑 Catalogo de Cuentas - {empresa}")
    st.write(f"Formato de fecha: {formato_fecha} | Moneda: {moneda}")

    # INICIALIZAR LISTA DE CUENTAS SI NO EXISTE
    if "catalogo" not in st.session_state:
        st.session_state["catalogo"] = []

    # FORMULARIO PARA AGREGAR CUENTA
    with st.form("form_catalogo"):
        col1, col2 = st.columns([1, 3])
        with col1:
            codigo = st.text_input("Codigo de Cuenta", placeholder="Ej: 101")
        with col2:
            nombre = st.text_input("Nombre de la Cuenta", placeholder="Ej: Caja General")
        agregar = st.form_submit_button("Agregar Cuenta")

    # VALIDAR Y AGREGAR
    if agregar:
        errores = []
        if not codigo.strip().isdigit():
            errores.append("El codigo debe ser numerico.")
        if not nombre.strip():
            errores.append("El nombre de la cuenta no puede estar vacio.")

        existe = any(c["codigo"] == codigo for c in st.session_state["catalogo"])
        if existe:
            errores.append("Ya existe una cuenta con ese codigo.")

        if errores:
            st.error("⚠️ Errores:\n- " + "\n- ".join(errores))
        else:
            st.session_state["catalogo"].append({
                "codigo": codigo,
                "nombre": nombre
            })
            st.success("✅ Cuenta agregada.")

    # MOSTRAR CATALOGO ACTUAL
    if st.session_state["catalogo"]:
        st.markdown("### 📋 Lista de Cuentas Registradas")

        for i, cuenta in enumerate(st.session_state["catalogo"]):
            col1, col2, col3 = st.columns([1, 4, 1])
            with col1:
                st.write(cuenta["codigo"])
            with col2:
                st.write(cuenta["nombre"])
            with col3:
                if st.button("🗑️", key=f"eliminar_cuenta_{i}"):
                    st.session_state["catalogo"].pop(i)
                    st.rerun()

        # EXPORTAR A EXCEL
        df_cat = pd.DataFrame(st.session_state["catalogo"])
        output_cat = BytesIO()
        with pd.ExcelWriter(output_cat, engine='openpyxl') as writer:
            df_cat.to_excel(writer, index=False, sheet_name='Catalogo')
        output_cat.seek(0)
        st.download_button(
            label="📁 Exportar Catalogo a Excel",
            data=output_cat,
            file_name="catalogo_cuentas.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        # FUNCION PARA GENERAR PDF DEL CATALOGO
        def generar_pdf_catalogo(catalogo, empresa):
            buffer = BytesIO()
            c = canvas.Canvas(buffer, pagesize=letter)
            c.setFont("Helvetica", 10)

            ancho, alto = letter
            margen = 40
            y = alto - 50

            c.drawString(margen, y, "MP-Systems - Catalogo de Cuentas")
            y -= 15
            c.drawString(margen, y, f"Empresa: {empresa}")
            y -= 15
            c.drawString(margen, y, f"Fecha del reporte: {datetime.now().strftime('%d/%m/%Y')}")
            y -= 25

            c.setFont("Helvetica-Bold", 9)
            c.drawString(margen, y, "Codigo")
            c.drawString(margen + 100, y, "Nombre de la Cuenta")
            y -= 15
            c.setFont("Helvetica", 9)

            for cuenta in catalogo:
                if y < 50:
                    c.showPage()
                    y = alto - 50
                c.drawString(margen, y, cuenta["codigo"])
                c.drawString(margen + 100, y, cuenta["nombre"][:60])
                y -= 15

            c.showPage()
            c.save()
            buffer.seek(0)
            return buffer

        # BOTON EXPORTAR A PDF
        pdf_catalogo = generar_pdf_catalogo(st.session_state["catalogo"], empresa)
        st.download_button(
            label="📄 Exportar Catalogo a PDF",
            data=pdf_catalogo,
            file_name="catalogo_cuentas.pdf",
            mime="application/pdf"
        )

    else:
        st.info("No hay cuentas registradas aun.")






# Modulo: ESTADO DE RESULTADOS
elif seleccion == "Estado de Resultados":
    st.subheader("📊 Estado de Resultados")
    empresa = st.session_state.get("empresa", "Empresa Desconocida")
    moneda = st.session_state.get("simbolo_moneda", "RD$")
    st.write(f"Empresa: {empresa}")

    if "estado_resultados" not in st.session_state:
        st.session_state["estado_resultados"] = []

    st.markdown("### ➕ Agregar Movimiento")

    tipo = st.selectbox("Tipo", ["Ingreso", "Gasto"])
    categoria = st.text_input("Categoria", placeholder="Ej: Ventas, Servicios, Alquiler")
    monto_str = st.text_input("Monto (Ej: RD$500.00)", placeholder="RD$0.00")

    try:
        monto = float(monto_str.replace(moneda, "").replace(",", "").strip())
    except:
        monto = 0.0

    if st.button("Agregar"):
        if not categoria.strip():
            st.warning("La categoria no puede estar vacia.")
        elif monto <= 0:
            st.warning("El monto debe ser mayor que 0.")
        else:
            nuevo = {"tipo": tipo, "categoria": categoria.strip(), "monto": monto}
            st.session_state["estado_resultados"].append(nuevo)
            st.success("Movimiento agregado correctamente.")

    if st.session_state["estado_resultados"]:
        st.markdown("### 🧾 Asientos Registrados")

        total_ingresos = sum(item["monto"] for item in st.session_state["estado_resultados"] if item["tipo"] == "Ingreso")
        total_gastos = sum(item["monto"] for item in st.session_state["estado_resultados"] if item["tipo"] == "Gasto")
        utilidad_neta = total_ingresos - total_gastos

        st.write(f"**Total Ingresos:** {moneda} {total_ingresos:,.2f}")
        st.write(f"**Total Gastos:** {moneda} {total_gastos:,.2f}")
        st.write(f"**Utilidad Neta:** {moneda} {utilidad_neta:,.2f}")

        df_resultados = pd.DataFrame(st.session_state["estado_resultados"])
        st.dataframe(df_resultados)

        # EXPORTAR A EXCEL CON FORMATO DE MONEDA
        output_excel = BytesIO()
        with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
            df_resultados.to_excel(writer, index=False, sheet_name='Estado de Resultados', startrow=0)

            wb = writer.book
            ws = writer.sheets['Estado de Resultados']

            col_monto = "C"
            col_tipo = "A"
            fila_inicio = 2
            fila_fin = ws.max_row
            fila_suma = fila_fin + 2

            ws[f"A{fila_suma}"] = "Total Ingresos:"
            ws[f"{col_monto}{fila_suma}"] = f'=SUMIF({col_tipo}{fila_inicio}:{col_tipo}{fila_fin},"Ingreso",{col_monto}{fila_inicio}:{col_monto}{fila_fin})'

            ws[f"A{fila_suma+1}"] = "Total Gastos:"
            ws[f"{col_monto}{fila_suma+1}"] = f'=SUMIF({col_tipo}{fila_inicio}:{col_tipo}{fila_fin},"Gasto",{col_monto}{fila_inicio}:{col_monto}{fila_fin})'

            ws[f"A{fila_suma+2}"] = "Utilidad Neta:"
            ws[f"{col_monto}{fila_suma+2}"] = f"={col_monto}{fila_suma}-{col_monto}{fila_suma+1}"

            ws[f"{col_monto}{fila_suma+2}"].font = Font(bold=True)

            if utilidad_neta >= 0:
                color = "C6EFCE"  # Verde claro
            else:
                color = "FFC7CE"  # Rojo claro

            ws[f"{col_monto}{fila_suma+2}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            ws[f"{col_monto}{fila_suma+2}"].number_format = f'"{moneda}" #,##0.00'

            # FORMATO DE MONEDA EN TODA LA COLUMNA DE MONTO
            for fila in range(2, fila_fin + 3):
                celda = ws[f"{col_monto}{fila}"]
                celda.number_format = f'"{moneda}" #,##0.00'

        output_excel.seek(0)
        st.download_button(
            label="📁 Exportar a Excel",
            data=output_excel,
            file_name="estado_resultados.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        # EXPORTAR A PDF
        def generar_pdf_resultado(data, empresa, moneda):
            buffer = BytesIO()
            c = canvas.Canvas(buffer, pagesize=letter)
            c.setFont("Helvetica", 10)
            ancho, alto = letter
            margen = 40
            y = alto - 50

            c.drawString(margen, y, "MP-Systems - Estado de Resultados")
            y -= 15
            c.drawString(margen, y, f"Empresa: {empresa}")
            y -= 15
            c.drawString(margen, y, f"Fecha del reporte: {datetime.now().strftime('%d/%m/%Y')}")
            y -= 25

            c.setFont("Helvetica-Bold", 9)
            c.drawString(margen, y, "Tipo")
            c.drawString(margen + 120, y, "Categoria")
            c.drawString(margen + 300, y, "Monto")
            y -= 15
            c.setFont("Helvetica", 9)

            for item in data:
                if y < 50:
                    c.showPage()
                    y = alto - 50
                c.drawString(margen, y, item["tipo"])
                c.drawString(margen + 120, y, item["categoria"][:40])
                c.drawString(margen + 300, y, f"{moneda} {item['monto']:,.2f}")
                y -= 15

            y -= 10
            c.setFont("Helvetica-Bold", 9)
            c.drawString(margen, y, f"Total Ingresos: {moneda} {total_ingresos:,.2f}")
            y -= 15
            c.drawString(margen, y, f"Total Gastos: {moneda} {total_gastos:,.2f}")
            y -= 15
            c.drawString(margen, y, f"Utilidad Neta: {moneda} {utilidad_neta:,.2f}")
            c.showPage()
            c.save()
            buffer.seek(0)
            return buffer

        pdf_resultado = generar_pdf_resultado(st.session_state["estado_resultados"], empresa, moneda)

        st.download_button(
            label="📄 Exportar a PDF",
            data=pdf_resultado,
            file_name="estado_resultados.pdf",
            mime="application/pdf"
        )

















# Modulo: BALANCE GENERAL
elif seleccion == "Balance General":
    st.subheader("🏦 Balance General")
    empresa = st.session_state.get("empresa", "Empresa Desconocida")
    moneda = st.session_state.get("simbolo_moneda", "RD$")
    st.write(f"Empresa: {empresa}")

    if "balance_general" not in st.session_state:
        st.session_state["balance_general"] = []

    st.markdown("### ➕ Agregar Cuenta")

    tipo = st.selectbox("Tipo", ["Activo", "Pasivo", "Patrimonio"])
    categoria = st.text_input("Cuenta / Categoria", placeholder="Ej: Caja, Proveedores, Capital")
    monto_str = st.text_input("Monto (Ej: RD$5,000.00)", placeholder="RD$0.00")

    try:
        monto = float(monto_str.replace(moneda, "").replace(",", "").strip())
    except:
        monto = 0.0

    if st.button("Agregar"):
        if not categoria.strip():
            st.warning("El campo 'Cuenta' no puede estar vacio.")
        elif monto <= 0:
            st.warning("El monto debe ser mayor que 0.")
        else:
            nuevo = {"tipo": tipo, "cuenta": categoria.strip(), "monto": monto}
            st.session_state["balance_general"].append(nuevo)
            st.success("Cuenta agregada correctamente.")

    if st.session_state["balance_general"]:
        st.markdown("### 📋 Balance Registrado")

        total_activo = sum(item["monto"] for item in st.session_state["balance_general"] if item["tipo"] == "Activo")
        total_pasivo = sum(item["monto"] for item in st.session_state["balance_general"] if item["tipo"] == "Pasivo")
        total_patrimonio = sum(item["monto"] for item in st.session_state["balance_general"] if item["tipo"] == "Patrimonio")
        resultado = total_activo - (total_pasivo + total_patrimonio)

        st.write(f"**Total Activos:** {moneda} {total_activo:,.2f}")
        st.write(f"**Total Pasivos:** {moneda} {total_pasivo:,.2f}")
        st.write(f"**Total Patrimonio:** {moneda} {total_patrimonio:,.2f}")
        st.write(f"**Diferencia:** {moneda} {resultado:,.2f}")

        df_balance = pd.DataFrame(st.session_state["balance_general"])
        st.dataframe(df_balance)

        # EXPORTAR A EXCEL
        output_excel = BytesIO()
        with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
            df_balance.to_excel(writer, index=False, sheet_name='Balance General', startrow=0)

            ws = writer.sheets['Balance General']
            col_monto = "C"
            col_tipo = "A"
            fila_inicio = 2
            fila_fin = ws.max_row
            fila_suma = fila_fin + 2

            ws[f"A{fila_suma}"] = "Total Activo:"
            ws[f"{col_monto}{fila_suma}"] = f'=SUMIF({col_tipo}{fila_inicio}:{col_tipo}{fila_fin},"Activo",{col_monto}{fila_inicio}:{col_monto}{fila_fin})'

            ws[f"A{fila_suma+1}"] = "Total Pasivo:"
            ws[f"{col_monto}{fila_suma+1}"] = f'=SUMIF({col_tipo}{fila_inicio}:{col_tipo}{fila_fin},"Pasivo",{col_monto}{fila_inicio}:{col_monto}{fila_fin})'

            ws[f"A{fila_suma+2}"] = "Total Patrimonio:"
            ws[f"{col_monto}{fila_suma+2}"] = f'=SUMIF({col_tipo}{fila_inicio}:{col_tipo}{fila_fin},"Patrimonio",{col_monto}{fila_inicio}:{col_monto}{fila_fin})'

            ws[f"A{fila_suma+3}"] = "Diferencia:"
            ws[f"{col_monto}{fila_suma+3}"] = f"={col_monto}{fila_suma}-{col_monto}{fila_suma+1}-{col_monto}{fila_suma+2}"

            if resultado == 0:
                color = "C6EFCE"  # Verde
            else:
                color = "FFC7CE"  # Rojo

            ws[f"{col_monto}{fila_suma+3}"].font = Font(bold=True)
            ws[f"{col_monto}{fila_suma+3}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            ws[f"{col_monto}{fila_suma+3}"].number_format = f'"{moneda}" #,##0.00'

            for fila in range(2, fila_fin + 4):
                celda = ws[f"{col_monto}{fila}"]
                celda.number_format = f'"{moneda}" #,##0.00'

        output_excel.seek(0)
        st.download_button(
            label="📁 Exportar a Excel",
            data=output_excel,
            file_name="balance_general.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        # EXPORTAR A PDF
        def generar_pdf_balance(data, empresa, moneda):
            buffer = BytesIO()
            c = canvas.Canvas(buffer, pagesize=letter)
            c.setFont("Helvetica", 10)
            ancho, alto = letter
            margen = 40
            y = alto - 50

            c.drawString(margen, y, "MP-Systems - Balance General")
            y -= 15
            c.drawString(margen, y, f"Empresa: {empresa}")
            y -= 15
            c.drawString(margen, y, f"Fecha del reporte: {datetime.now().strftime('%d/%m/%Y')}")
            y -= 25

            c.setFont("Helvetica-Bold", 9)
            c.drawString(margen, y, "Tipo")
            c.drawString(margen + 120, y, "Cuenta")
            c.drawString(margen + 300, y, "Monto")
            y -= 15
            c.setFont("Helvetica", 9)

            for item in data:
                if y < 50:
                    c.showPage()
                    y = alto - 50
                c.drawString(margen, y, item["tipo"])
                c.drawString(margen + 120, y, item["cuenta"][:40])
                c.drawString(margen + 300, y, f"{moneda} {item['monto']:,.2f}")
                y -= 15

            y -= 10
            c.setFont("Helvetica-Bold", 9)
            c.drawString(margen, y, f"Total Activo: {moneda} {total_activo:,.2f}")
            y -= 15
            c.drawString(margen, y, f"Total Pasivo: {moneda} {total_pasivo:,.2f}")
            y -= 15
            c.drawString(margen, y, f"Total Patrimonio: {moneda} {total_patrimonio:,.2f}")
            y -= 15
            c.drawString(margen, y, f"Diferencia: {moneda} {resultado:,.2f}")
            c.showPage()
            c.save()
            buffer.seek(0)
            return buffer

        pdf_balance = generar_pdf_balance(st.session_state["balance_general"], empresa, moneda)

        st.download_button(
            label="📄 Exportar a PDF",
            data=pdf_balance,
            file_name="balance_general.pdf",
            mime="application/pdf"
        )



# Modulo: HISTORIAL DEL SISTEMA
elif seleccion == "Historial del Sistema":
    st.subheader("📜 Historial del Sistema")
    empresa = st.session_state.get("empresa", "Empresa Desconocida")
    st.write(f"Empresa: {empresa}")

    if "historial" not in st.session_state:
        st.session_state["historial"] = []

    # FORMULARIO PARA AGREGAR REGISTRO AL HISTORIAL
    st.markdown("### ➕ Agregar Registro")
    modulo = st.selectbox("Modulo", ["Libro Diario", "Catálogo", "Balance", "Estado de Resultados", "Otros"])
    descripcion = st.text_input("Descripcion de la accion", placeholder="Ej: Se registraron 3 asientos.")
    fecha = st.date_input("Fecha", datetime.today())

    if st.button("Registrar Historial"):
        if not descripcion.strip():
            st.warning("Debe ingresar una descripcion valida.")
        else:
            nuevo = {
                "fecha": fecha.strftime("%d/%m/%Y"),
                "modulo": modulo,
                "descripcion": descripcion.strip()
            }
            st.session_state["historial"].append(nuevo)
            st.success("Registro agregado correctamente.")

    if st.session_state["historial"]:
        st.markdown("### 📂 Historial Registrado")
        df_historial = pd.DataFrame(st.session_state["historial"])
        st.dataframe(df_historial)

        # EXPORTAR A EXCEL
        output_excel = BytesIO()
        with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
            df_historial.to_excel(writer, index=False, sheet_name='Historial')
        output_excel.seek(0)

        st.download_button(
            label="📁 Exportar a Excel",
            data=output_excel,
            file_name="historial_sistema.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        # EXPORTAR A PDF
        def generar_pdf_historial(data, empresa):
            buffer = BytesIO()
            c = canvas.Canvas(buffer, pagesize=letter)
            c.setFont("Helvetica", 10)
            ancho, alto = letter
            margen = 40
            y = alto - 50

            c.drawString(margen, y, "MP-Systems - Historial del Sistema")
            y -= 15
            c.drawString(margen, y, f"Empresa: {empresa}")
            y -= 15
            c.drawString(margen, y, f"Fecha del reporte: {datetime.now().strftime('%d/%m/%Y')}")
            y -= 25

            c.setFont("Helvetica-Bold", 9)
            c.drawString(margen, y, "Fecha")
            c.drawString(margen + 80, y, "Modulo")
            c.drawString(margen + 200, y, "Descripcion")
            y -= 15
            c.setFont("Helvetica", 9)

            for item in data:
                if y < 50:
                    c.showPage()
                    y = alto - 50
                c.drawString(margen, y, item["fecha"])
                c.drawString(margen + 80, y, item["modulo"])
                c.drawString(margen + 200, y, item["descripcion"][:60])
                y -= 15

            c.showPage()
            c.save()
            buffer.seek(0)
            return buffer

        pdf_historial = generar_pdf_historial(st.session_state["historial"], empresa)

        st.download_button(
            label="📄 Exportar a PDF",
            data=pdf_historial,
            file_name="historial_sistema.pdf",
            mime="application/pdf"
        )


